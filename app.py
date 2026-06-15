import base64
import csv
import io
import json
import os
import pickle
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

import fitz  # PyMuPDF
import numpy as np
from PIL import Image
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# Shared helpers
# -----------------------------

APP_DIR = Path(__file__).resolve().parent
LOGO_PATH = APP_DIR / "white_line_trucking_logo.png"
DATA_DIR = APP_DIR / "saved_reports"
GITHUB_STORAGE_DEFAULT_PATH = "portal_storage"


def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _report_path(report_name):
    safe = re.sub(r"[^A-Za-z0-9_-]", "_", str(report_name))
    return DATA_DIR / f"{safe}.pkl"


def _secret_or_env(name, default=""):
    """Read a Streamlit secret first, then an environment variable."""
    try:
        value = st.secrets.get(name, "")
        if value:
            return str(value)
    except Exception:
        pass
    return str(os.environ.get(name, default) or default)


def _github_storage_config():
    """Return GitHub repo storage config from Streamlit secrets/env vars.

    Required secrets:
      GITHUB_TOKEN = GitHub fine-grained token with Contents read/write for the repo
      GITHUB_REPO = owner/repository-name
    Optional:
      GITHUB_BRANCH = main
      GITHUB_STORAGE_PATH = portal_storage
    """
    token = _secret_or_env("GITHUB_TOKEN")
    repo = _secret_or_env("GITHUB_REPO")
    branch = _secret_or_env("GITHUB_BRANCH", "main")
    storage_path = _secret_or_env("GITHUB_STORAGE_PATH", GITHUB_STORAGE_DEFAULT_PATH).strip("/")
    if token and repo:
        return {"token": token, "repo": repo, "branch": branch, "storage_path": storage_path}
    return None


def _github_request(method, url, token, payload=None):
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Accept", "application/vnd.github+json")
    req.add_header("X-GitHub-Api-Version", "2022-11-28")
    if payload is not None:
        req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=20) as response:
        body = response.read().decode("utf-8")
        return json.loads(body) if body else {}


def _github_report_file(report_name):
    safe = re.sub(r"[^A-Za-z0-9_-]", "_", str(report_name))
    cfg = _github_storage_config()
    if not cfg:
        return None, None
    file_path = f"{cfg['storage_path']}/{safe}.pkl"
    return cfg, file_path


def _load_report_from_github(report_name):
    cfg, file_path = _github_report_file(report_name)
    if not cfg:
        return None
    encoded_path = "/".join(urllib.parse.quote(part) for part in file_path.split("/"))
    url = f"https://api.github.com/repos/{cfg['repo']}/contents/{encoded_path}?ref={urllib.parse.quote(cfg['branch'])}"
    try:
        obj = _github_request("GET", url, cfg["token"])
        content = obj.get("content", "")
        if not content:
            return None
        raw = base64.b64decode(content.encode("utf-8"))
        return pickle.loads(raw)
    except urllib.error.HTTPError as exc:
        if exc.code == 404:
            return None
        return None
    except Exception:
        return None


def _save_report_to_github(report_name, payload):
    cfg, file_path = _github_report_file(report_name)
    if not cfg:
        return False
    encoded_path = "/".join(urllib.parse.quote(part) for part in file_path.split("/"))
    base_url = f"https://api.github.com/repos/{cfg['repo']}/contents/{encoded_path}"
    sha = None
    try:
        current = _github_request("GET", f"{base_url}?ref={urllib.parse.quote(cfg['branch'])}", cfg["token"])
        sha = current.get("sha")
    except urllib.error.HTTPError as exc:
        if exc.code != 404:
            return False
    except Exception:
        return False

    raw = pickle.dumps(payload)
    body = {
        "message": f"Update portal saved report: {report_name}",
        "content": base64.b64encode(raw).decode("utf-8"),
        "branch": cfg["branch"],
    }
    if sha:
        body["sha"] = sha

    try:
        _github_request("PUT", base_url, cfg["token"], body)
        return True
    except Exception:
        return False


def save_persistent_report(report_name, payload):
    """Save the latest processed report.

    Local disk is used for local/LAN testing. GitHub repo storage is used on Streamlit
    Cloud when GITHUB_TOKEN and GITHUB_REPO are configured in Streamlit secrets.
    """
    try:
        _ensure_data_dir()
        path = _report_path(report_name)
        tmp_path = path.with_suffix(".tmp")
        with tmp_path.open("wb") as f:
            pickle.dump(payload, f)
        tmp_path.replace(path)
    except Exception:
        pass

    # This is the permanent cloud copy. It survives Streamlit Cloud sleep/reboot/rebuilds.
    _save_report_to_github(report_name, payload)


def load_persistent_report(report_name):
    # Prefer the GitHub-backed copy so Streamlit Cloud reboot/reset does not erase reports.
    saved = _load_report_from_github(report_name)
    if saved is not None:
        return saved

    # Local fallback for testing on a computer or office network.
    try:
        path = _report_path(report_name)
        if path.exists():
            with path.open("rb") as f:
                return pickle.load(f)
    except Exception:
        return None
    return None


def permanent_storage_is_configured():
    return _github_storage_config() is not None


def initialize_persistent_report_state():
    """Reload saved reports into Streamlit session state when the app/browser reopens."""
    if "cash_forecast_data" not in st.session_state:
        saved = load_persistent_report("cash_forecast")
        if saved:
            st.session_state["cash_forecast_data"] = saved.get("data")
            st.session_state["cash_forecast_filename"] = saved.get("filename", "saved report")

    if "ar_report_data" not in st.session_state:
        saved = load_persistent_report("ar_report")
        if saved:
            st.session_state["ar_report_data"] = saved

    if "ap_report_data" not in st.session_state:
        saved = load_persistent_report("ap_report")
        if saved:
            st.session_state["ap_report_data"] = saved

    for upload_key, report_name in [("pl_cm_upload", "pl_cm_report"), ("pl_ytd_upload", "pl_ytd_report")]:
        state_key = f"{upload_key}_matrix"
        filename_key = f"{upload_key}_filename"
        if state_key not in st.session_state:
            saved = load_persistent_report(report_name)
            if saved:
                st.session_state[state_key] = saved.get("matrix")
                st.session_state[filename_key] = saved.get("filename", "saved report")


def money_to_float(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    try:
        number = float(text)
        return -number if neg else number
    except ValueError:
        return 0.0


def format_money_no_symbol(value):
    return f"{money_to_float(value):,.2f}"


def calculate_promise_date(invoice_date):
    """Return invoice date + 30 days in M/D/YYYY format."""
    raw = str(invoice_date or "").strip()
    if not raw:
        return ""
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(raw, fmt) + timedelta(days=30)
            return f"{dt.month}/{dt.day}/{dt.year}"
        except ValueError:
            continue
    return ""


def ar_output_row(row):
    invoice_date = row.get("Date", "")
    return {
        "WO #": row.get("WO #", ""),
        "": "",
        "Date": invoice_date,
        "Type": "Invoice",
        "Total": row.get("Total", ""),
        "Promise date": row.get("Promise date") or calculate_promise_date(invoice_date),
    }


def format_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    # openpyxl may return date objects without datetime
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.month}/{value.day}/{value.year}"
    text = str(value).strip()
    if not text:
        return ""
    # Excel/QBO sometimes stores datetimes as text.
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            d = datetime.strptime(text, fmt)
            return f"{d.month}/{d.day}/{d.year}"
        except ValueError:
            pass
    return text


def build_tsv(rows, columns, include_header=False, include_total=False):
    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    if include_header:
        writer.writerow(columns)
    for r in rows:
        writer.writerow([r.get(c, "") for c in columns])
    if include_total:
        total = sum(money_to_float(r.get("Open balance", "")) for r in rows)
        total_row = [""] * len(columns)
        if columns:
            total_row[0] = "TOTAL"
        if "Open balance" in columns:
            total_row[columns.index("Open balance")] = format_money_no_symbol(total)
        writer.writerow(total_row)
    return output.getvalue()


def render_copy_button(label, text_to_copy, key_suffix):
    js_text = json.dumps(text_to_copy)
    safe_id = re.sub(r"[^A-Za-z0-9_-]", "_", key_suffix)
    components.html(
        f"""
        <div style="font-family: Arial, sans-serif; margin: 0 0 12px 0;">
            <button id="copy-btn-{safe_id}" style="
                background:#E11919;
                color:white;
                border:none;
                border-radius:6px;
                padding:10px 16px;
                cursor:pointer;
                font-size:14px;
                font-weight:700;">
                {label}
            </button>
            <span id="copy-status-{safe_id}" style="margin-left:10px; font-size:14px;"></span>
        </div>
        <script>
        const copyText_{safe_id} = {js_text};
        const button_{safe_id} = document.getElementById("copy-btn-{safe_id}");
        const status_{safe_id} = document.getElementById("copy-status-{safe_id}");

        async function copyToClipboard_{safe_id}() {{
            try {{
                if (navigator.clipboard && window.isSecureContext) {{
                    await navigator.clipboard.writeText(copyText_{safe_id});
                }} else {{
                    const textArea = document.createElement("textarea");
                    textArea.value = copyText_{safe_id};
                    textArea.style.position = "fixed";
                    textArea.style.left = "-999999px";
                    textArea.style.top = "-999999px";
                    document.body.appendChild(textArea);
                    textArea.focus();
                    textArea.select();
                    document.execCommand("copy");
                    textArea.remove();
                }}
                status_{safe_id}.textContent = "Copied. Paste directly into Excel.";
                status_{safe_id}.style.color = "#0b6b0b";
            }} catch (err) {{
                status_{safe_id}.textContent = "Copy failed. Use the manual copy box below.";
                status_{safe_id}.style.color = "#b00020";
            }}
        }}
        button_{safe_id}.addEventListener("click", copyToClipboard_{safe_id});
        </script>
        """,
        height=58,
    )


def render_ap_two_copy_buttons(no_due_text, due_date_text):
    """Render the two A/P copy buttons as one visible component."""
    no_due_json = json.dumps(no_due_text)
    due_json = json.dumps(due_date_text)
    components.html(
        f"""
        <div style="font-family: Arial, sans-serif; margin-top: 10px; padding: 16px; border: 1px solid #d7d7d7; border-radius: 12px; background: #ffffff;">
            <div style="font-size: 16px; font-weight: 800; margin-bottom: 12px; color: #111;">A/P Copy Buttons</div>
            <button id="copy-no-due" style="
                background:#E11919;
                color:white;
                border:none;
                border-radius:8px;
                padding:12px 18px;
                cursor:pointer;
                font-size:14px;
                font-weight:800;
                margin-right:12px;
                margin-bottom:10px;">
                Copy all cells except Due Date
            </button>
            <button id="copy-due-only" style="
                background:#111111;
                color:white;
                border:none;
                border-radius:8px;
                padding:12px 18px;
                cursor:pointer;
                font-size:14px;
                font-weight:800;
                margin-bottom:10px;">
                Copy Due Date column only
            </button>
            <div id="copy-ap-status" style="font-size:14px; margin-top:6px; min-height:20px;"></div>
        </div>
        <script>
        const noDueText = {no_due_json};
        const dueDateText = {due_json};
        const status = document.getElementById("copy-ap-status");

        async function copyText(text, message) {{
            try {{
                if (navigator.clipboard) {{
                    await navigator.clipboard.writeText(text);
                }} else {{
                    const textArea = document.createElement("textarea");
                    textArea.value = text;
                    textArea.style.position = "fixed";
                    textArea.style.left = "-999999px";
                    textArea.style.top = "-999999px";
                    document.body.appendChild(textArea);
                    textArea.focus();
                    textArea.select();
                    document.execCommand("copy");
                    textArea.remove();
                }}
                status.textContent = message;
                status.style.color = "#0b6b0b";
            }} catch (err) {{
                status.textContent = "Copy failed. Click once inside the app window and try again.";
                status.style.color = "#b00020";
            }}
        }}

        document.getElementById("copy-no-due").addEventListener("click", function() {{
            copyText(noDueText, "Copied all A/P cells except Due Date. Paste directly into Excel.");
        }});
        document.getElementById("copy-due-only").addEventListener("click", function() {{
            copyText(dueDateText, "Copied Due Date column only. Paste directly into Excel.");
        }});
        </script>
        """,
        height=132,
    )


def render_ar_two_copy_buttons(no_promise_text, promise_date_text):
    """Render the two A/R copy buttons as one visible component."""
    no_promise_json = json.dumps(no_promise_text)
    promise_json = json.dumps(promise_date_text)
    components.html(
        f"""
        <div style="font-family: Arial, sans-serif; margin-top: 10px; padding: 16px; border: 1px solid #d7d7d7; border-radius: 12px; background: #ffffff;">
            <div style="font-size: 16px; font-weight: 800; margin-bottom: 12px; color: #111;">A/R Copy Buttons</div>
            <button id="copy-ar-no-promise" style="
                background:#E11919;
                color:white;
                border:none;
                border-radius:8px;
                padding:12px 18px;
                cursor:pointer;
                font-size:14px;
                font-weight:800;
                margin-right:12px;
                margin-bottom:10px;">
                Copy all cells except Promise date
            </button>
            <button id="copy-ar-promise-only" style="
                background:#111111;
                color:white;
                border:none;
                border-radius:8px;
                padding:12px 18px;
                cursor:pointer;
                font-size:14px;
                font-weight:800;
                margin-bottom:10px;">
                Copy Promise date column only
            </button>
            <div id="copy-ar-status" style="font-size:14px; margin-top:6px; min-height:20px;"></div>
        </div>
        <script>
        const arNoPromiseText = {no_promise_json};
        const arPromiseText = {promise_json};
        const arStatus = document.getElementById("copy-ar-status");

        async function copyArText(text, message) {{
            try {{
                if (navigator.clipboard) {{
                    await navigator.clipboard.writeText(text);
                }} else {{
                    const textArea = document.createElement("textarea");
                    textArea.value = text;
                    textArea.style.position = "fixed";
                    textArea.style.left = "-999999px";
                    textArea.style.top = "-999999px";
                    document.body.appendChild(textArea);
                    textArea.focus();
                    textArea.select();
                    document.execCommand("copy");
                    textArea.remove();
                }}
                arStatus.textContent = message;
                arStatus.style.color = "#0b6b0b";
            }} catch (err) {{
                arStatus.textContent = "Copy failed. Click once inside the app window and try again.";
                arStatus.style.color = "#b00020";
            }}
        }}

        document.getElementById("copy-ar-no-promise").addEventListener("click", function() {{
            copyArText(arNoPromiseText, "Copied all A/R cells except Promise date. Paste directly into Excel.");
        }});
        document.getElementById("copy-ar-promise-only").addEventListener("click", function() {{
            copyArText(arPromiseText, "Copied Promise date column only. Paste directly into Excel.");
        }});
        </script>
        """,
        height=132,
    )


# -----------------------------
# A/R completed work order PDF extraction
# -----------------------------


def extract_completed_orders(pdf_bytes, scale=3, checked_threshold=180):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    completed = []
    all_rows = []

    for page_index, page in enumerate(doc, start=1):
        words = page.get_text("words")
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        arr = np.array(img)

        mask = (
            (arr[:, :, 2] > 80)
            & (arr[:, :, 0] < 170)
            & (arr[:, :, 1] < 180)
            & (arr[:, :, 2] > arr[:, :, 0] + 5)
            & (arr[:, :, 2] > arr[:, :, 1] - 25)
        )

        candidates = []
        for w in words:
            x0, y0, x1, y1, text, *_ = w
            if 30 <= x0 <= 55 and 120 <= y0 <= 760 and re.fullmatch(r"\d{3,4}", text):
                candidates.append({"wo": text, "y0": float(y0), "y1": float(y1)})

        candidates = sorted(candidates, key=lambda r: r["y0"])

        for i, row_start in enumerate(candidates):
            next_y = candidates[i + 1]["y0"] if i + 1 < len(candidates) else 760
            y_start = row_start["y0"] - 2
            y_end = min(next_y - 1, row_start["y0"] + 45)

            date_words = []
            for w in words:
                x0, y0, x1, y1, text, *_ = w
                cy = (y0 + y1) / 2
                if 65 <= x0 <= 132 and y_start <= cy <= y_end:
                    date_words.append((float(y0), float(x0), text))
            date_text = "".join(t for _, __, t in sorted(date_words))
            date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", date_text)
            date_value = date_match.group(1) if date_match else date_text

            money_words = []
            for w in words:
                x0, y0, x1, y1, text, *_ = w
                cy = (y0 + y1) / 2
                if x0 >= 510 and y_start <= cy <= y_end and re.fullmatch(r"\$[\d,]+\.\d{2}", text):
                    money_words.append((float(y0), float(x0), text))
            total_value = money_words[0][2] if money_words else ""

            x1p = int(151 * scale)
            x2p = int(160 * scale)
            y1p = int((row_start["y0"] + 4) * scale)
            y2p = int((row_start["y0"] + 13) * scale)
            blue_count = int(mask[y1p:y2p, x1p:x2p].sum())
            is_complete = blue_count > checked_threshold

            row = {
                "WO #": row_start["wo"],
                "": "",
                "Date": date_value,
                "Type": "Invoice",
                "Total": total_value,
                "Promise date": calculate_promise_date(date_value),
                "_Page": page_index,
                "_Complete": is_complete,
                "_CheckScore": blue_count,
            }
            all_rows.append(row)
            if is_complete:
                completed.append(row)

    return completed, all_rows


def col_letter(n):
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def cell_xml(row_num, col_num, value, style=None, is_number=False):
    ref = f"{col_letter(col_num)}{row_num}"
    style_attr = f' s="{style}"' if style is not None else ""
    if value is None:
        return f'<c r="{ref}"{style_attr}/>'
    if is_number:
        return f'<c r="{ref}"{style_attr}><v>{value}</v></c>'
    text = xml_escape(str(value))
    return f'<c r="{ref}" t="inlineStr"{style_attr}><is><t>{text}</t></is></c>'


def build_ar_xlsx(rows):
    sheet_rows = []
    sheet_rows.append([("WO #", 1, False), ("", 1, False), ("Date", 1, False), ("Type", 1, False), ("Total", 1, False), ("Promise date", 1, False)])

    for r in rows:
        output_row = ar_output_row(r)
        amount = money_to_float(output_row.get("Total", ""))
        sheet_rows.append([
            (output_row.get("WO #", ""), None, False),
            ("", None, False),
            (output_row.get("Date", ""), None, False),
            ("Invoice", None, False),
            (f"{amount:.2f}", 2, True),
            (output_row.get("Promise date", ""), None, False),
        ])

    row_xml = []
    for r_idx, row in enumerate(sheet_rows, start=1):
        cells = []
        for c_idx, (value, style, is_number) in enumerate(row, start=1):
            cells.append(cell_xml(r_idx, c_idx, value, style=style, is_number=is_number))
        row_xml.append(f'<row r="{r_idx}">{"".join(cells)}</row>')

    sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
           xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <cols>
    <col min="1" max="1" width="12" customWidth="1"/>
    <col min="2" max="2" width="8" customWidth="1"/>
    <col min="3" max="3" width="14" customWidth="1"/>
    <col min="4" max="4" width="14" customWidth="1"/>
    <col min="5" max="5" width="14" customWidth="1"/>
    <col min="6" max="6" width="16" customWidth="1"/>
  </cols>
  <sheetData>
    {''.join(row_xml)}
  </sheetData>
</worksheet>'''

    workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Output" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>'''

    workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''

    root_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>'''

    styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <numFmts count="1"><numFmt numFmtId="164" formatCode="$#,##0.00"/></numFmts>
  <fonts count="2">
    <font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/><family val="2"/></font>
  </fonts>
  <fills count="3">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFE11919"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center"/></xf>
    <xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''

    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>'''

    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    core_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
 xmlns:dc="http://purl.org/dc/elements/1.1/"
 xmlns:dcterms="http://purl.org/dc/terms/"
 xmlns:dcmitype="http://purl.org/dc/dcmitype/"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>White Line Trucking Management Portal</dc:creator>
  <cp:lastModifiedBy>White Line Trucking Management Portal</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>
</cp:coreProperties>'''

    app_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
 xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>White Line Trucking Management Portal</Application>
</Properties>'''

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", root_rels)
        z.writestr("xl/workbook.xml", workbook_xml)
        z.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        z.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        z.writestr("xl/styles.xml", styles_xml)
        z.writestr("docProps/core.xml", core_xml)
        z.writestr("docProps/app.xml", app_xml)
    return output.getvalue()


def build_ar_csv(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["WO #", "", "Date", "Type", "Total", "Promise date"])
    for r in rows:
        output_row = ar_output_row(r)
        writer.writerow([output_row.get("WO #", ""), "", output_row.get("Date", ""), "Invoice", output_row.get("Total", ""), output_row.get("Promise date", "")])
    return output.getvalue().encode("utf-8-sig")


# -----------------------------
# A/P Aging Detail parsing and output
# -----------------------------

AP_COLUMNS = ["Vendor display name", "Num", "Date", "Transaction type", "Open balance", "Due date"]
AP_COPY_NO_DUE_COLUMNS = ["Vendor display name", "Num", "Date", "Transaction type", "Open balance"]
AP_COPY_DUE_DATE_COLUMNS = ["Due date"]


def normalize_header(text):
    return re.sub(r"[^a-z0-9]", "", str(text or "").strip().lower())


# QuickBooks exports can vary slightly by browser/version and may include
# sort arrows, line breaks, hidden columns, or shortened header labels.
# Keep these keys exactly aligned with AP_COLUMNS.
HEADER_ALIASES = {
    "Vendor display name": {"vendordisplayname", "vendor", "vendorname", "name", "supplier", "payee"},
    "Num": {"num", "no", "number", "billno", "billnumber", "docnum", "documentnumber", "refno", "reference"},
    "Date": {"date", "transactiondate", "txndate", "txn date", "transdate"},
    "Transaction type": {"transactiontype", "type", "transaction", "transtype"},
    "Open balance": {"openbalance", "openbal", "balance", "amountopen", "amountdue", "openamount", "remainingbalance", "open"},
    "Due date": {"duedate", "due", "billduedate", "duedate"},
}


def find_header_map(row):
    normalized = [normalize_header(c) for c in row]
    header_map = {}
    for target, aliases in HEADER_ALIASES.items():
        normalized_aliases = {normalize_header(a) for a in aliases}
        for idx, cell in enumerate(normalized):
            if not cell:
                continue
            if cell in normalized_aliases:
                header_map[target] = idx
                break
            # Helpful for QuickBooks headers with extra sort text/icons, such as
            # "Vendor display name sort ascending" or "Open balance $".
            if any(alias and (cell.startswith(alias) or alias in cell) for alias in normalized_aliases):
                header_map[target] = idx
                break
    if "Vendor display name" in header_map and "Open balance" in header_map:
        return header_map
    return None


def clean_vendor_name(value):
    text = str(value or "").strip()
    # Remove common indentation symbols from QBO group rows/copy output.
    text = re.sub(r"^[\u25be\u25b8\u25bc\u25ba\s]+", "", text).strip()
    return text


def should_skip_ap_row(row):
    vendor = str(row.get("Vendor display name", "")).strip()
    tx_type = str(row.get("Transaction type", "")).strip()
    if not vendor:
        return True
    lower_vendor = vendor.lower()
    if lower_vendor.startswith("out of range"):
        return True
    if lower_vendor.startswith("total for"):
        return True
    if lower_vendor == "total":
        return True
    if "total" in lower_vendor and not tx_type:
        return True
    # Keep actual vendor bill rows. Most QBO A/P Aging Detail rows have a transaction type.
    if not tx_type and not row.get("Open balance"):
        return True
    return False


def coerce_ap_row(raw):
    vendor = clean_vendor_name(raw.get("Vendor display name", ""))
    num = "" if raw.get("Num") is None else str(raw.get("Num", "")).strip()
    tx_type = "" if raw.get("Transaction type") is None else str(raw.get("Transaction type", "")).strip()
    open_balance = format_money_no_symbol(raw.get("Open balance", ""))
    return {
        "Vendor display name": vendor,
        "Num": num,
        "Date": format_date(raw.get("Date", "")),
        "Transaction type": tx_type,
        "Open balance": open_balance,
        "Due date": format_date(raw.get("Due date", "")),
    }


def extract_ap_from_xlsx(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    # QuickBooks exports usually put the report on the first sheet, but not always.
    # Scan all worksheets and use the first one that contains the A/P header row.
    all_rows = None
    for ws_candidate in wb.worksheets:
        candidate_rows = [[cell for cell in row] for row in ws_candidate.iter_rows(values_only=True)]
        if any(find_header_map(row) for row in candidate_rows):
            all_rows = candidate_rows
            break
    if all_rows is None:
        ws = wb.active
        all_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    company_name = "White Line Trucking LLC"
    report_name = "A/P Aging Detail Report"
    as_of = ""

    for row in all_rows[:12]:
        joined = " ".join(str(c).strip() for c in row if c is not None and str(c).strip())
        if not joined:
            continue
        if "a/p aging" in joined.lower() or "ap aging" in joined.lower():
            report_name = joined
        if joined.lower().startswith("as of"):
            as_of = joined
        elif "as of" in joined.lower():
            match = re.search(r"as of\s+(.+)", joined, flags=re.IGNORECASE)
            if match:
                as_of = "As of " + match.group(1).strip()
        elif not any(x in joined.lower() for x in ["quickbooks", "report", "date"]):
            # First useful centered title row is usually company name.
            if company_name == "White Line Trucking LLC" and ("white" in joined.lower() or "line" in joined.lower()):
                company_name = joined

    header_idx = None
    header_map = None
    for idx, row in enumerate(all_rows):
        possible_map = find_header_map(row)
        if possible_map:
            header_idx = idx
            header_map = possible_map
            break

    if header_idx is None or header_map is None:
        raise ValueError("Could not find the A/P Aging Detail column header row. Export the report to Excel from QuickBooks and upload that file.")

    cleaned = []
    for raw_row in all_rows[header_idx + 1:]:
        raw = {}
        for col in AP_COLUMNS:
            idx = header_map.get(col)
            raw[col] = raw_row[idx] if idx is not None and idx < len(raw_row) else ""
        row = coerce_ap_row(raw)
        if not should_skip_ap_row(row):
            cleaned.append(row)

    if not as_of:
        today = datetime.today()
        as_of = f"As of {today.strftime('%B')} {today.day}, {today.year}"

    return cleaned, {"company": company_name, "report": report_name, "as_of": as_of}



def parse_ap_csv_text(text):
    rows_raw = list(csv.reader(io.StringIO(text)))
    company = "White Line Trucking LLC"
    report = "A/P Aging Detail Report"
    as_of = ""

    for row in rows_raw[:12]:
        joined = " ".join(str(c).strip() for c in row if c and str(c).strip())
        if not joined:
            continue
        if "a/p aging" in joined.lower() or "ap aging" in joined.lower():
            report = joined
        if joined.lower().startswith("as of"):
            as_of = joined
        if ("white" in joined.lower() or "line" in joined.lower()) and len([c for c in row if c]) <= 3:
            company = joined

    header_idx = None
    header_map = None
    for idx, row in enumerate(rows_raw):
        possible_map = find_header_map(row)
        if possible_map:
            header_idx = idx
            header_map = possible_map
            break

    if header_idx is None or header_map is None:
        # Fall back to pasted-text parser if the file is not a normal CSV table.
        return parse_ap_pasted_text(text)

    rows = []
    for raw_row in rows_raw[header_idx + 1:]:
        raw = {}
        for col in AP_COLUMNS:
            idx = header_map.get(col)
            raw[col] = raw_row[idx] if idx is not None and idx < len(raw_row) else ""
        row = coerce_ap_row(raw)
        if not should_skip_ap_row(row):
            rows.append(row)

    return rows, {"company": company, "report": report, "as_of": as_of}

def parse_ap_pasted_text(text):
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return [], {"company": "White Line Trucking LLC", "report": "A/P Aging Detail Report", "as_of": ""}

    company = "White Line Trucking LLC"
    report = "A/P Aging Detail Report"
    as_of = ""
    header_idx = None
    header_map = None
    split_rows = []

    for ln in lines:
        if "\t" in ln:
            parts = [p.strip() for p in ln.split("\t")]
        else:
            parts = [p.strip() for p in re.split(r"\s{2,}", ln) if p.strip()]
        split_rows.append(parts)

    for i, parts in enumerate(split_rows):
        joined = " ".join(parts)
        if "a/p aging" in joined.lower() or "ap aging" in joined.lower():
            report = joined
        if joined.lower().startswith("as of"):
            as_of = joined
        if ("white" in joined.lower() or "line" in joined.lower()) and len(parts) <= 3:
            company = joined
        possible_map = find_header_map(parts)
        if possible_map:
            header_idx = i
            header_map = possible_map
            break

    if header_idx is None:
        # Fallback for copied lines from the visible table without headers. Assumes columns match screenshot order.
        header_map = {c: i for i, c in enumerate(AP_COLUMNS)}
        header_idx = -1

    rows = []
    for parts in split_rows[header_idx + 1:]:
        if len(parts) < 3:
            continue
        raw = {}
        for col in AP_COLUMNS:
            idx = header_map.get(col)
            raw[col] = parts[idx] if idx is not None and idx < len(parts) else ""
        row = coerce_ap_row(raw)
        if not should_skip_ap_row(row):
            rows.append(row)

    return rows, {"company": company, "report": report, "as_of": as_of}


def build_ap_xlsx(rows, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "AP Aging Detail"

    red = "E11919"
    black = "111111"
    gray = "EDEDED"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin_gray)

    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws.merge_cells("A3:F3")
    ws["A1"] = meta.get("company", "White Line Trucking LLC")
    ws["A2"] = meta.get("report", "A/P Aging Detail Report")
    ws["A3"] = meta.get("as_of", "")
    for cell in ("A1", "A2", "A3"):
        ws[cell].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(bold=True, size=16, color=black)
    ws["A2"].font = Font(size=11, color=black)
    ws["A3"].font = Font(size=11, color=black)

    header_row = 5
    for idx, col in enumerate(AP_COLUMNS, start=1):
        c = ws.cell(row=header_row, column=idx, value=col)
        c.font = Font(bold=True, color=black)
        c.fill = PatternFill("solid", fgColor=gray)
        c.border = border
        c.alignment = Alignment(horizontal="left")

    start_row = header_row + 1
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, col in enumerate(AP_COLUMNS, start=1):
            value = row.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == "Open balance" else "left")
            if col == "Open balance":
                cell.value = money_to_float(value)
                cell.number_format = '#,##0.00'

    total_row = start_row + len(rows)
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 5, sum(money_to_float(r.get("Open balance")) for r in rows))
    ws.cell(total_row, 5).number_format = '#,##0.00'
    for c_idx in range(1, 7):
        cell = ws.cell(total_row, c_idx)
        cell.font = Font(bold=True, color=black)
        cell.border = Border(top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))

    widths = [28, 14, 13, 18, 16, 13]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A6"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_ap_csv(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(AP_COLUMNS)
    for r in rows:
        writer.writerow([r.get(c, "") for c in AP_COLUMNS])
    total = [""] * len(AP_COLUMNS)
    total[0] = "TOTAL"
    total[4] = format_money_no_symbol(sum(money_to_float(r.get("Open balance")) for r in rows))
    writer.writerow(total)
    return output.getvalue().encode("utf-8-sig")


def render_ap_report(rows, meta):
    total = sum(money_to_float(r.get("Open balance", "")) for r in rows)
    headers = AP_COLUMNS
    html_rows = []
    for r in rows:
        html_rows.append(
            "<tr>" + "".join(
                f'<td class="money">{r.get(h, "")}</td>' if h == "Open balance" else f"<td>{xml_escape(str(r.get(h, '')))}</td>"
                for h in headers
            ) + "</tr>"
        )
    total_row = "".join([
        '<td class="total-label">TOTAL</td>',
        '<td></td>', '<td></td>', '<td></td>',
        f'<td class="money total-money">{format_money_no_symbol(total)}</td>',
        '<td></td>'
    ])

    table_html = f"""
    <div class="ap-report-card">
        <div class="ap-title">{xml_escape(meta.get('company', 'White Line Trucking LLC'))}</div>
        <div class="ap-subtitle">{xml_escape(meta.get('report', 'A/P Aging Detail Report'))}</div>
        <div class="ap-asof">{xml_escape(meta.get('as_of', ''))}</div>
        <table class="ap-table">
            <thead><tr>{''.join(f'<th>{h}</th>' for h in headers)}</tr></thead>
            <tbody>{''.join(html_rows)}<tr class="total-row">{total_row}</tr></tbody>
        </table>
    </div>
    """
    st.markdown(table_html, unsafe_allow_html=True)


def demo_ap_rows():
    rows = [
        {"Vendor display name": "Cleburne Ford", "Num": "", "Date": "5/30/2026", "Transaction type": "Bill", "Open balance": "5,774.58", "Due date": "6/30/2026"},
        {"Vendor display name": "FleetPride", "Num": "", "Date": "5/31/2026", "Transaction type": "Bill", "Open balance": "7,826.39", "Due date": "6/30/2026"},
        {"Vendor display name": "FleetPride", "Num": "", "Date": "5/31/2026", "Transaction type": "Bill", "Open balance": "4,424.96", "Due date": "6/30/2026"},
        {"Vendor display name": "O'Reilly Auto Parts", "Num": "", "Date": "5/28/2026", "Transaction type": "Bill", "Open balance": "2,809.41", "Due date": "6/20/2026"},
        {"Vendor display name": "Texas State Comptroller", "Num": "", "Date": "6/2/2026", "Transaction type": "Bill", "Open balance": "5,471.70", "Due date": "6/20/2026"},
        {"Vendor display name": "Premier Truck Group", "Num": "", "Date": "5/31/2026", "Transaction type": "Bill", "Open balance": "3,910.52", "Due date": "6/10/2026"},
        {"Vendor display name": "Gillam, Wharram & Co., P.C.", "Num": "", "Date": "5/12/2026", "Transaction type": "Bill", "Open balance": "7,845.00", "Due date": "5/12/2026"},
    ]
    meta = {"company": "White Line Trucking LLC", "report": "A/P Aging Detail Report", "as_of": "As of June 10, 2026"}
    return rows, meta




# -----------------------------
# Cash Forecast workbook reader
# -----------------------------

CASH_SHEET_NAME = "Cash Flow & Cash Mgt"
CASH_WEEK_COLS = list(range(5, 18))  # 13-week current rolling quarter, columns E:Q


def excel_serial_to_display_date(value, short_year=False):
    """Convert Excel serial dates or Python date/datetime values to display text."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{str(value.year)[-2:] if short_year else value.year}"
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.month}/{value.day}/{str(value.year)[-2:] if short_year else value.year}"
    if isinstance(value, (int, float)):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=float(value))
            return f"{d.month}/{d.day}/{str(d.year)[-2:] if short_year else d.year}"
        except Exception:
            return str(value)
    return str(value).strip()


def cash_number(value, decimals=0):
    if value is None or value == "":
        return ""
    if isinstance(value, str) and value.strip() == "":
        return ""
    try:
        number = float(value)
    except Exception:
        return str(value)
    if abs(number) < 0.005:
        number = 0
    if decimals == 0:
        return f"{number:,.0f}"
    return f"{number:,.{decimals}f}"


def cash_value_to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def load_cash_sheet_values(file_bytes, filename):
    """Return cell values for the Cash Flow & Cash Mgt sheet as a dict keyed by (row, col)."""
    name = (filename or "").lower()
    values = {}

    if name.endswith(".xlsb"):
        from pyxlsb import open_workbook
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsb") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            with open_workbook(tmp_path) as wb:
                sheet_name = CASH_SHEET_NAME if CASH_SHEET_NAME in wb.sheets else None
                if sheet_name is None:
                    for s in wb.sheets:
                        if "cash" in s.lower() and "flow" in s.lower():
                            sheet_name = s
                            break
                if sheet_name is None:
                    raise ValueError("Could not find the Cash Flow & Cash Mgt sheet in this workbook.")
                sh = wb.get_sheet(sheet_name)
                for r_idx, row in enumerate(sh.rows(), start=1):
                    if r_idx > 340:
                        break
                    for c_idx, cell in enumerate(row, start=1):
                        if c_idx > 18:
                            break
                        values[(r_idx, c_idx)] = cell.v
        finally:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass
    elif name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_name = CASH_SHEET_NAME if CASH_SHEET_NAME in wb.sheetnames else None
        if sheet_name is None:
            for s in wb.sheetnames:
                if "cash" in s.lower() and "flow" in s.lower():
                    sheet_name = s
                    break
        if sheet_name is None:
            raise ValueError("Could not find the Cash Flow & Cash Mgt sheet in this workbook.")
        ws = wb[sheet_name]
        for r in range(1, 341):
            for c in range(1, 19):
                values[(r, c)] = ws.cell(r, c).value
    else:
        raise ValueError("Upload the Cash Forecast workbook as .xlsb, .xlsx, or .xlsm.")

    return values


def extract_cash_forecast(file_bytes, filename):
    values = load_cash_sheet_values(file_bytes, filename)
    getv = lambda r, c: values.get((r, c), "")

    company = str(getv(1, 2) or "White Line Trucking LLC").strip()
    title = str(getv(1, 5) or "Cash Forecast Model").strip()
    owner_line = str(getv(3, 5) or "").strip()
    weeks = [excel_serial_to_display_date(getv(4, c), short_year=True) for c in CASH_WEEK_COLS]

    definitions = [
        ("section", "(A) BEGINNING BALANCE CASH:", None),
        ("data", "Beginning Balance Cash", 5),
        ("data", "Balance - LOC", 6),
        ("data", "Used from LOC", 7),
        ("data", "Paid to LOC", 8),
        ("section", "(B) CASH RECEIPTS (ESTIMATED)", None),
        ("data", "Received on A/R (AR Forecast)", 10),
        ("total", "TOTAL INFLOW:", 111),
        ("section", "(C) CASH DISBURSEMENTS (ESTIMATED)", None),
        ("data", "Payrolls", 113),
        ("data", "Recurring Expenses - (Recurring Tab)", 114),
        ("data", "Accts Payable (AP Forecast)", 115),
        ("total", "TOTAL OUTFLOW:", 216),
        ("section", "(D) INVESTMENT IMPACT (ESTIMATED)", None),
        ("data", "Project Investment - Consulting", 219),
        ("data", "Project Investment - TAX", 220),
        ("data", "Project Investment - TAX Implementation", 221),
        ("data", "Project Investment - Valuation", 222),
        ("data", "Project Investment - Recruiting Services", 223),
        ("data", "Funding (+ Net Proceeds)", 224),
        ("data", "Funding (- Payment)", 225),
        ("data", "Annualized Bloodwall Benefit Sub Total", 327),
        ("total", "TOTAL INVESTMENT IMPACT FLOW:", 328),
        ("section", "(E) ENDING CASH AVAILABLE", None),
        ("total", "ENDING CASH AVAILABLE - BEFORE SWEEP:", 331),
        ("total", "ENDING CASH AVAILABLE - AFTER SWEEP:", 332),
    ]

    rows = []
    for kind, label, source_row in definitions:
        if source_row is None:
            rows.append({"kind": kind, "label": label, "values": [""] * len(CASH_WEEK_COLS)})
        else:
            raw_values = [getv(source_row, c) for c in CASH_WEEK_COLS]
            rows.append({"kind": kind, "label": label, "values": raw_values})

    ending_before = [cash_value_to_float(getv(331, c)) for c in CASH_WEEK_COLS]
    ending_before_nonblank = [v for v in ending_before if v is not None]
    metrics = {
        "first_week_ending": getv(331, CASH_WEEK_COLS[0]),
        "quarter_end_cash": getv(331, CASH_WEEK_COLS[-1]),
        "lowest_cash": min(ending_before_nonblank) if ending_before_nonblank else None,
        "actual_adgl": getv(333, 5),
        "confidence": getv(333, 8),
        "four_weeks_float": getv(333, 11),
    }

    return {
        "company": company,
        "title": title,
        "owner_line": owner_line,
        "weeks": weeks,
        "rows": rows,
        "metrics": metrics,
    }


def render_cash_forecast(data):
    weeks = data["weeks"]
    header_cells = ''.join(f'<th class="week-col">{xml_escape(str(w))}</th>' for w in weeks)
    body_rows = []
    for row in data["rows"]:
        kind = row["kind"]
        label = xml_escape(str(row["label"]))
        if kind == "section":
            body_rows.append(f'<tr class="cash-section"><td colspan="{len(weeks) + 1}">{label}</td></tr>')
        else:
            cells = ''.join(f'<td class="num">{cash_number(v)}</td>' for v in row["values"])
            row_class = "cash-total" if kind == "total" else "cash-data"
            body_rows.append(f'<tr class="{row_class}"><td class="cash-label">{label}</td>{cells}</tr>')

    html = f"""
        <div class="cash-card">
            <div class="cash-title">{xml_escape(data.get('company', 'White Line Trucking LLC'))}</div>
            <div class="cash-subtitle">{xml_escape(data.get('title', 'Cash Forecast Model'))}</div>
            <div class="cash-owner">{xml_escape(data.get('owner_line', ''))}</div>
            <div class="cash-scroll">
                <table class="cash-table">
                    <thead><tr><th class="label-col">Cash Forecast Line</th>{header_cells}</tr></thead>
                    <tbody>{''.join(body_rows)}</tbody>
                </table>
            </div>
        </div>
        """
    st.markdown(html, unsafe_allow_html=True)


# -----------------------------
# Styling and layout
# -----------------------------

st.set_page_config(
    page_title="White Line Trucking Management Portal",
    page_icon="💎",
    layout="wide",
)


def apply_diamond_theme():
    st.markdown(
        """
        <style>
        :root {
            --diamond-red: #D0A400;
            --diamond-red-dark: #8A6D00;
            --diamond-black: #070707;
            --diamond-charcoal: #181818;
            --diamond-gray: #F6F6F6;
            --diamond-border: #D7D7D7;
        }
        #MainMenu, footer {visibility: hidden;}
        .stApp {
            background: linear-gradient(180deg, #ffffff 0%, #f7f7f7 58%, #eeeeee 100%);
        }
        .block-container {
            padding-top: 1.2rem;
            max-width: 1220px;
        }
        .diamond-header {
            background: linear-gradient(135deg, #050505 0%, #171717 55%, #2B0000 100%);
            border-left: 8px solid var(--diamond-red);
            border-radius: 18px;
            padding: 22px 26px;
            margin-bottom: 18px;
            box-shadow: 0 12px 28px rgba(0,0,0,.18);
        }
        .portal-title {
            color: white;
            font-size: 1.55rem;
            font-weight: 800;
            margin-top: 10px;
            letter-spacing: .02em;
        }
        .portal-subtitle {
            color: #d6d6d6;
            margin-top: 4px;
            font-size: .98rem;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid var(--diamond-border);
            border-top: 4px solid var(--diamond-red);
            border-radius: 14px;
            padding: 14px 16px;
            box-shadow: 0 4px 14px rgba(0,0,0,.06);
        }
        div[data-testid="stMetric"] *,
        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] *,
        [data-testid="stMetricValue"],
        [data-testid="stMetricValue"] * {
            color: #111111 !important;
            opacity: 1 !important;
        }
        [data-testid="stFileUploader"] label,
        [data-testid="stFileUploader"] p,
        [data-testid="stFileUploader"] span {
            color: #111111 !important;
            opacity: 1 !important;
        }
        .block-container h1, .block-container h2, .block-container h3,
        .block-container h4, .block-container h5, .block-container h6,
        .block-container .stMarkdown p,
        .block-container label {
            color: #111111;
        }
        .diamond-header .portal-title, .diamond-header .portal-title *,
        .home-hero .home-title, .home-hero .home-title * {
            color: #ffffff !important;
        }
        .diamond-header .portal-subtitle, .diamond-header .portal-subtitle *,
        .home-hero .home-subtitle, .home-hero .home-subtitle * {
            color: #e6e6e6 !important;
        }
        .home-hero .home-note, .home-hero .home-note * {
            color: #ffffff !important;
        }
        section[data-testid="stSidebar"] {
            background: #0b0b0b;
        }
        section[data-testid="stSidebar"] * {
            color: #ffffff;
        }
        .stDownloadButton > button, .stButton > button {
            background: var(--diamond-red);
            color: #ffffff;
            border: 1px solid var(--diamond-red-dark);
            border-radius: 9px;
            font-weight: 700;
        }
        .stDownloadButton > button:hover, .stButton > button:hover {
            background: var(--diamond-red-dark);
            color: #ffffff;
            border-color: var(--diamond-red-dark);
        }
        [data-testid="stFileUploader"] {
            background: #ffffff !important;
            border: 1px solid var(--diamond-border) !important;
            border-radius: 14px !important;
            padding: 10px 14px !important;
        }
        [data-testid="stFileUploader"] section {
            background: #f6f7f9 !important;
            border: 2px dashed #b8b8b8 !important;
            border-radius: 10px !important;
        }
        [data-testid="stFileUploader"] section * {
            color: #111111 !important;
        }
        [data-testid="stFileUploader"] small,
        [data-testid="stFileUploader"] p,
        [data-testid="stFileUploader"] span,
        [data-testid="stFileUploader"] label {
            color: #111111 !important;
            opacity: 1 !important;
        }
        [data-testid="stFileUploader"] button {
            background: #ffffff !important;
            color: #111111 !important;
            border: 1px solid #b8b8b8 !important;
            border-radius: 8px !important;
            font-weight: 700 !important;
        }
        [data-testid="stFileUploader"] button:hover {
            background: #E11919 !important;
            color: #ffffff !important;
            border-color: #B00000 !important;
        }
        .small-muted {
            color: #666;
            font-size: .9rem;
        }
        .ap-report-card {
            background: white;
            border: 1px solid #d4d4d4;
            border-radius: 8px;
            padding: 20px 24px 24px 24px;
            box-shadow: 0 4px 16px rgba(0,0,0,.06);
            margin-top: 10px;
        }
        .ap-title {
            text-align: center;
            font-size: 22px;
            font-weight: 800;
            color: #222;
            line-height: 1.2;
        }
        .ap-subtitle, .ap-asof {
            text-align: center;
            color: #222;
            font-size: 14px;
            line-height: 1.35;
        }
        table.ap-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 14px;
            color: #1d1d1d;
        }
        .ap-table th {
            text-align: left;
            font-weight: 800;
            border-top: 1px solid #d0d0d0;
            border-bottom: 2px solid #c8c8c8;
            padding: 8px 8px;
            white-space: nowrap;
        }
        .ap-table td {
            padding: 6px 8px;
            border-bottom: 1px solid #eeeeee;
            vertical-align: top;
        }
        .ap-table td.money, .ap-table th:nth-child(5) {
            text-align: right;
        }
        .ap-table tr.total-row td {
            border-top: 1px solid #bcbcbc;
            border-bottom: 1px solid #bcbcbc;
            font-weight: 800;
            padding-top: 8px;
            padding-bottom: 8px;
        }
        .total-money {
            font-weight: 800;
        }
        .total-label {
            font-weight: 800;
        }

        .cash-card {
            background: #ffffff;
            border: 1px solid #d4d4d4;
            border-radius: 10px;
            padding: 18px 20px 22px 20px;
            box-shadow: 0 4px 16px rgba(0,0,0,.06);
            margin-top: 12px;
        }
        .cash-title {
            text-align: center;
            font-size: 22px;
            font-weight: 800;
            color: #111;
        }
        .cash-subtitle, .cash-owner {
            text-align: center;
            font-size: 14px;
            color: #222;
            line-height: 1.35;
        }
        .cash-scroll {
            width: 100%;
            overflow-x: auto;
            margin-top: 18px;
            border: 1px solid #cfcfcf;
            border-radius: 8px;
        }
        table.cash-table {
            width: max-content;
            min-width: 100%;
            border-collapse: collapse;
            font-size: 12px;
            color: #111;
        }
        .cash-table th {
            background: #0070C0;
            color: #ffffff;
            border: 1px solid #2c5f8d;
            padding: 6px 8px;
            white-space: nowrap;
            font-weight: 800;
            text-align: center;
        }
        .cash-table .label-col {
            min-width: 275px;
            text-align: left;
            position: sticky;
            left: 0;
            z-index: 2;
        }
        .cash-table .week-col {
            min-width: 82px;
        }
        .cash-table td {
            border: 1px solid #d9d9d9;
            padding: 5px 7px;
            white-space: nowrap;
        }
        .cash-table td.cash-label {
            font-weight: 600;
            min-width: 275px;
            background: #f7f7f7;
            position: sticky;
            left: 0;
        }
        .cash-table td.num {
            text-align: right;
            min-width: 82px;
        }
        .cash-section td {
            background: #d9eaf7;
            color: #111;
            font-weight: 900;
            border-top: 2px solid #111;
        }
        .cash-total td {
            background: #eeeeee;
            font-weight: 900;
            border-top: 2px solid #777;
        }
        .cash-total td.cash-label {
            background: #e6e6e6;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_brand_header(show_home_button=True):
    logo_html = ""
    if LOGO_PATH.exists():
        logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
        logo_img = f'<img src="data:image/png;base64,{logo_b64}" alt="White Line Trucking LLC" style="max-width:340px;width:100%;height:auto;" />'
        # The logo is also a Home link. The sidebar reads the page query string before the navigation widget is created.
        logo_html = f'<a href="?page=Home" target="_self" title="Return to Home">{logo_img}</a>'
    st.markdown(
        f'''
        <div class="diamond-header">
            {logo_html}
            <div class="portal-title">Management Portal</div>
            <div class="portal-subtitle">Cash Flow, A/R, A/P, and P&L reporting tools for White Line Trucking LLC</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )
    if show_home_button:
        if st.button("🏠 Home", key=f"home_shortcut_{st.session_state.get('main_page_nav', 'page')}"):
            st.session_state["main_page_nav_target"] = "Home"
            st.rerun()


def render_summary_cards(cards):
    """Render three clean, native summary boxes for report pages."""
    if not cards:
        return

    visible_cards = [(str(label), str(value), str(note) if note else "") for label, value, note in cards]
    columns = st.columns(len(visible_cards))

    for col, (label, value, note) in zip(columns, visible_cards):
        with col:
            try:
                with st.container(border=True):
                    st.metric(label=label, value=value)
                    if note:
                        st.caption(note)
            except TypeError:
                st.metric(label=label, value=value)
                if note:
                    st.caption(note)


def get_requested_page_from_query():
    try:
        page = st.query_params.get("page", None)
        if isinstance(page, list):
            page = page[0] if page else None
        return page
    except Exception:
        try:
            params = st.experimental_get_query_params()
            vals = params.get("page", [])
            return vals[0] if vals else None
        except Exception:
            return None


def clear_page_query_param():
    try:
        if "page" in st.query_params:
            del st.query_params["page"]
    except Exception:
        try:
            st.experimental_set_query_params()
        except Exception:
            pass


apply_diamond_theme()

# Extra styling for the home/title page and left navigation
st.markdown(
    """
    <style>
    .home-hero {
        background: linear-gradient(135deg, #050505 0%, #151515 55%, #2B0000 100%);
        border-left: 10px solid #E11919;
        border-radius: 22px;
        padding: 34px 34px 30px 34px;
        margin: 8px 0 22px 0;
        box-shadow: 0 14px 34px rgba(0,0,0,.20);
    }
    .home-hero img {
        max-width: 380px;
        width: 100%;
        height: auto;
        margin-bottom: 18px;
    }
    .home-title {
        color: #ffffff;
        font-size: 2.15rem;
        font-weight: 900;
        letter-spacing: .02em;
        margin: 0;
    }
    .home-subtitle {
        color: #e6e6e6;
        font-size: 1.08rem;
        margin-top: 8px;
        max-width: 920px;
        line-height: 1.45;
    }
    .home-note {
        color: #ffffff;
        background: rgba(225,25,25,.20);
        border: 1px solid rgba(225,25,25,.55);
        border-radius: 12px;
        padding: 12px 14px;
        margin-top: 18px;
        display: inline-block;
        font-weight: 700;
    }
    .home-card {
        background: #ffffff;
        border: 1px solid #d7d7d7;
        border-top: 5px solid #E11919;
        border-radius: 16px;
        padding: 18px 18px 16px 18px;
        min-height: 178px;
        box-shadow: 0 6px 18px rgba(0,0,0,.07);
    }
    .home-card-title {
        font-size: 1.08rem;
        font-weight: 900;
        color: #111111;
        margin-bottom: 8px;
    }
    .home-card-text {
        color: #444444;
        font-size: .96rem;
        line-height: 1.42;
    }
    .section-title {
        font-size: 1.30rem;
        font-weight: 900;
        color: #111111;
        margin: 16px 0 10px 0;
    }
    .summary-strip {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        gap: 12px;
        margin: 14px 0 18px 0;
    }
    .summary-card {
        background: #ffffff !important;
        border: 1px solid #d8d8d8 !important;
        border-left: 6px solid #E11919 !important;
        border-radius: 14px !important;
        padding: 14px 16px !important;
        box-shadow: 0 4px 14px rgba(0,0,0,.06) !important;
    }
    .summary-label {
        color: #555555 !important;
        opacity: 1 !important;
        font-size: .82rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: .03em;
    }
    .summary-value {
        color: #111111 !important;
        opacity: 1 !important;
        font-size: 1.45rem;
        font-weight: 900;
        margin-top: 3px;
        line-height: 1.15;
    }
    .summary-note {
        color: #777777 !important;
        opacity: 1 !important;
        font-size: .82rem;
        margin-top: 3px;
        min-height: 16px;
    }
    section[data-testid="stSidebar"] .stRadio label {
        font-weight: 800;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label {
        background: rgba(255,255,255,.07);
        border: 1px solid rgba(255,255,255,.12);
        border-radius: 10px;
        padding: 8px 10px;
        margin-bottom: 6px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def trim_pl_matrix(matrix):
    # Remove fully blank outer rows/columns while preserving the report layout inside.
    def is_blank_row(row):
        return all(str(c).strip() == "" for c in row)

    rows = [list(r) for r in matrix]
    while rows and is_blank_row(rows[0]):
        rows.pop(0)
    while rows and is_blank_row(rows[-1]):
        rows.pop()
    if not rows:
        return []

    max_len = max(len(r) for r in rows)
    rows = [r + [""] * (max_len - len(r)) for r in rows]

    non_blank_cols = []
    for col_idx in range(max_len):
        if any(str(row[col_idx]).strip() != "" for row in rows):
            non_blank_cols.append(col_idx)
    if not non_blank_cols:
        return []

    start_col = min(non_blank_cols)
    end_col = max(non_blank_cols)
    return [row[start_col:end_col + 1] for row in rows]


def format_pl_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.month}/{value.day}/{value.year}"
    if isinstance(value, float):
        if abs(value - round(value)) < 0.000001:
            return f"{int(round(value)):,}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value).strip()


def extract_pl_from_xlsx(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrix = []
    for row in ws.iter_rows():
        matrix.append([format_pl_cell(cell.value) for cell in row])
    return trim_pl_matrix(matrix)


def extract_pl_from_csv_text(text):
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    matrix = [[format_pl_cell(c) for c in row] for row in reader]
    return trim_pl_matrix(matrix)


def render_pl_report(matrix, title):
    if not matrix:
        st.warning("No report rows were found in the uploaded file.")
        return

    rows_html = []
    for r_idx, row in enumerate(matrix):
        row_text = " ".join(str(c).strip().lower() for c in row if str(c).strip())
        non_blank = [str(c).strip() for c in row if str(c).strip()]
        cls = ""
        if r_idx <= 2:
            cls = "pl-title-row"
        elif len(non_blank) == 1 and not re.search(r"\d", non_blank[0]):
            cls = "pl-section-row"
        elif row_text.startswith("total") or " net income" in f" {row_text}" or row_text in {"gross profit", "net operating income"}:
            cls = "pl-total-row"

        cells = []
        for c_idx, cell in enumerate(row):
            text = xml_escape(str(cell))
            cell_cls = "pl-label" if c_idx == 0 else "pl-num"
            if cls:
                cell_cls += f" {cls}"
            cells.append(f'<td class="{cell_cls}">{text}</td>')
        rows_html.append(f'<tr>{"".join(cells)}</tr>')

    html = f"""
    <style>
    .pl-wrap {{
        overflow-x: auto;
        border: 1px solid #d7d7d7;
        border-radius: 12px;
        background: #ffffff;
        margin-top: 14px;
        padding: 8px;
    }}
    .pl-table {{
        border-collapse: collapse;
        width: max-content;
        min-width: 100%;
        font-family: Arial, sans-serif;
        font-size: 13px;
        color: #111111;
    }}
    .pl-table td {{
        border-bottom: 1px solid #ececec;
        padding: 6px 10px;
        white-space: nowrap;
        vertical-align: middle;
    }}
    .pl-table td.pl-label {{
        min-width: 260px;
        text-align: left;
        font-weight: 500;
        position: sticky;
        left: 0;
        background: #ffffff;
        z-index: 1;
    }}
    .pl-table td.pl-num {{
        min-width: 120px;
        text-align: right;
    }}
    .pl-title-row {{
        font-weight: 900 !important;
        text-align: center !important;
        background: #f5f5f5 !important;
        font-size: 14px;
    }}
    .pl-section-row {{
        font-weight: 900 !important;
        background: #fafafa !important;
        color: #111111;
    }}
    .pl-total-row {{
        font-weight: 900 !important;
        border-top: 2px solid #111111 !important;
        background: #f1f1f1 !important;
    }}
    </style>
    <div class="pl-wrap">
        <table class="pl-table">
            {''.join(rows_html)}
        </table>
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


def render_pl_page(page_title, upload_key):
    render_brand_header()
    st.subheader(page_title)
    st.caption("Upload the QuickBooks Profit & Loss Excel export. This page is view-only and does not include copy or download buttons.")

    pl_upload = st.file_uploader("Upload QuickBooks Profit & Loss export", type=["xlsx", "xlsm", "csv", "txt"], key=upload_key)
    state_key = f"{upload_key}_matrix"
    filename_key = f"{upload_key}_filename"

    if pl_upload is not None:
        file_bytes = pl_upload.getvalue()
        name = pl_upload.name.lower()
        try:
            if name.endswith((".xlsx", ".xlsm")):
                matrix = extract_pl_from_xlsx(file_bytes)
            else:
                text = file_bytes.decode("utf-8-sig", errors="replace")
                matrix = extract_pl_from_csv_text(text)
            st.session_state[state_key] = matrix
            st.session_state[filename_key] = pl_upload.name
            report_name = "pl_cm_report" if upload_key == "pl_cm_upload" else "pl_ytd_report"
            save_persistent_report(report_name, {"matrix": matrix, "filename": pl_upload.name, "saved_at": datetime.now().isoformat()})
            st.success("P&L report loaded successfully.")
        except Exception as exc:
            st.error("The P&L file could not be processed. Export the Profit & Loss report from QuickBooks to Excel and upload that file.")
            st.exception(exc)

    matrix = st.session_state.get(state_key)
    if matrix is not None:
        if pl_upload is None:
            st.info(f"Showing last uploaded {page_title} report: {st.session_state.get(filename_key, 'saved report')}")
        render_pl_report(matrix, page_title)
    elif pl_upload is None:
        st.info("Upload the QuickBooks Profit & Loss export to view it here.")


def render_sidebar_nav():
    pages = ["Home", "Cash Flow", "A/R", "A/P", "P&L CM", "P&L YTD"]

    # Home buttons and clickable logo set a target page. Apply it BEFORE the sidebar radio is created.
    # Streamlit does not allow changing a widget key after the widget has been instantiated.
    query_page = get_requested_page_from_query()
    target_page = st.session_state.pop("main_page_nav_target", None) or query_page
    if target_page in pages:
        st.session_state["main_page_nav"] = target_page
        clear_page_query_param()

    if "main_page_nav" not in st.session_state or st.session_state["main_page_nav"] not in pages:
        st.session_state["main_page_nav"] = "Home"

    if LOGO_PATH.exists():
        logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
        st.sidebar.markdown(
            f'<a href="?page=Home" target="_self" title="Return to Home"><img src="data:image/png;base64,{logo_b64}" style="width:100%;height:auto;" /></a>',
            unsafe_allow_html=True,
        )
    st.sidebar.markdown("**White Line Trucking LLC**")
    st.sidebar.caption("Management Portal")
    st.sidebar.markdown("---")
    page = st.sidebar.radio(
        "Select Page",
        pages,
        key="main_page_nav",
    )
    st.sidebar.markdown("---")
    st.sidebar.caption("Click the logo or Home button to return to the title page.")
    return page


def render_home_page():
    logo_html = ""
    if LOGO_PATH.exists():
        logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
        logo_html = f'<img src="data:image/png;base64,{logo_b64}" alt="White Line Trucking LLC" />'

    st.markdown(
        f"""
        <div class="home-hero">
            {logo_html}
            <div class="home-title">White Line Trucking Management Portal</div>
            <div class="home-subtitle">
                One central place for managers to review weekly cash flow, A/R work orders, A/P aging detail, and P&L reports.
            </div>
            <div class="home-note">Use the left-side menu to select Cash Flow, A/R, A/P, P&L CM, or P&L YTD.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="section-title">Portal Pages</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            """
            <div class="home-card">
                <div class="home-card-title">Cash Flow</div>
                <div class="home-card-text">
                    Upload the updated Cash Forecast workbook and view the current rolling quarter in a clean manager-facing layout.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open Cash Flow", key="home_open_cash"):
            st.session_state["main_page_nav_target"] = "Cash Flow"
            st.rerun()
    with c2:
        st.markdown(
            """
            <div class="home-card">
                <div class="home-card-title">A/R</div>
                <div class="home-card-text">
                    Upload the Work In Progress PDF and pull only checked-complete work orders for invoice preparation.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open A/R", key="home_open_ar"):
            st.session_state["main_page_nav_target"] = "A/R"
            st.rerun()
    with c3:
        st.markdown(
            """
            <div class="home-card">
                <div class="home-card-title">A/P</div>
                <div class="home-card-text">
                    Upload the QuickBooks A/P Aging Detail export, remove grouping headers, and copy the payable data in the needed format.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open A/P", key="home_open_ap"):
            st.session_state["main_page_nav_target"] = "A/P"
            st.rerun()

    c4, c5 = st.columns(2)
    with c4:
        st.markdown(
            """
            <div class="home-card">
                <div class="home-card-title">P&L CM</div>
                <div class="home-card-text">
                    Upload the current-month QuickBooks Profit & Loss export and view it inside the portal.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open P&L CM", key="home_open_pl_cm"):
            st.session_state["main_page_nav_target"] = "P&L CM"
            st.rerun()
    with c5:
        st.markdown(
            """
            <div class="home-card">
                <div class="home-card-title">P&L YTD</div>
                <div class="home-card-text">
                    Upload the year-to-date QuickBooks Profit & Loss export and view it inside the portal.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open P&L YTD", key="home_open_pl_ytd"):
            st.session_state["main_page_nav_target"] = "P&L YTD"
            st.rerun()


def render_cash_flow_page():
    render_brand_header()
    st.subheader("Cash Flow")
    st.caption("Upload the updated Cash Forecast Model workbook. The portal reads the current rolling quarter from the Cash Flow & Cash Mgt sheet and displays a manager-view copy.")

    cash_upload = st.file_uploader("Upload Cash Forecast workbook", type=["xlsb", "xlsx", "xlsm"], key="cash_forecast_upload")

    if cash_upload is not None:
        file_bytes = cash_upload.getvalue()
        try:
            cash_data = extract_cash_forecast(file_bytes, cash_upload.name)
            st.session_state["cash_forecast_data"] = cash_data
            st.session_state["cash_forecast_filename"] = cash_upload.name
            save_persistent_report("cash_forecast", {"data": cash_data, "filename": cash_upload.name, "saved_at": datetime.now().isoformat()})
            st.success("Cash forecast loaded successfully.")
        except Exception as exc:
            st.error("The cash forecast workbook could not be processed. Upload the updated CFM workbook and make sure the Cash Flow & Cash Mgt sheet is present.")
            st.exception(exc)
            cash_data = st.session_state.get("cash_forecast_data")
    else:
        cash_data = st.session_state.get("cash_forecast_data")
        if cash_data is not None:
            st.info(f"Showing last uploaded Cash Flow report: {st.session_state.get('cash_forecast_filename', 'saved report')}")

    if cash_data is not None:
        metrics = cash_data.get("metrics", {})
        metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
        metric_col1.metric("1st Week Ending", f"${cash_number(metrics.get('first_week_ending'))}")
        metric_col2.metric("Quarter-End Cash", f"${cash_number(metrics.get('quarter_end_cash'))}")
        metric_col3.metric("Lowest Cash", f"${cash_number(metrics.get('lowest_cash'))}")
        metric_col4.metric("4 Weeks Float", f"${cash_number(metrics.get('four_weeks_float'))}")

        render_cash_forecast(cash_data)

        st.caption("Note: this is a read-only manager view. The bookkeeper should continue updating the Excel workbook, save it, and upload the updated file here.")
    elif cash_upload is None:
        st.info("Upload the updated CFM workbook to show the current rolling quarter page here. This page is designed to mirror the manager-facing cash forecast view without requiring managers to open the full Excel model.")


def render_ar_page():
    render_brand_header()
    st.subheader("A/R")
    st.caption("Upload the weekly Work In Progress List Report PDF to pull only checked-complete work orders.")

    uploaded = st.file_uploader("Upload weekly WIP PDF report", type=["pdf"], key="ar_pdf")

    if uploaded is not None:
        pdf_bytes = uploaded.getvalue()
        try:
            rows, all_rows = extract_completed_orders(pdf_bytes)
            ar_payload = {
                "rows": rows,
                "all_rows": all_rows,
                "filename": uploaded.name,
                "saved_at": datetime.now().isoformat(),
            }
            st.session_state["ar_report_data"] = ar_payload
            save_persistent_report("ar_report", ar_payload)
            st.success("Completed work orders extracted successfully.")
        except Exception as exc:
            st.error("The PDF could not be processed. Confirm it is the White Line Trucking Work In Progress List Report PDF.")
            st.exception(exc)

    ar_data = st.session_state.get("ar_report_data")

    if ar_data:
        rows = ar_data.get("rows", [])
        all_rows = ar_data.get("all_rows", [])
        if uploaded is None:
            st.info(f"Showing last uploaded A/R report: {ar_data.get('filename', 'saved report')}")
        total = sum(money_to_float(r.get("Total", "")) for r in rows)

        render_summary_cards([
            ("Completed WOs", f"{len(rows)}", "checked-complete only"),
            ("Total Dollars", f"${total:,.2f}", "invoice-ready value"),
            ("PDF Rows Reviewed", f"{len(all_rows)}", "audit count"),
        ])

        display_rows = [ar_output_row(r) for r in rows]
        st.dataframe(display_rows, use_container_width=True, hide_index=True)

        st.subheader("Copy A/R Output")
        ar_copy_no_promise_columns = ["WO #", "", "Date", "Type", "Total"]
        ar_copy_promise_columns = ["Promise date"]
        ar_tsv_without_promise = build_tsv(display_rows, ar_copy_no_promise_columns, include_header=False)
        ar_tsv_promise_only = build_tsv(display_rows, ar_copy_promise_columns, include_header=False)
        render_ar_two_copy_buttons(ar_tsv_without_promise, ar_tsv_promise_only)

        with st.expander("A/R audit details"):
            st.write(f"Total PDF rows reviewed: {len(all_rows)}")
            audit_rows = [{"WO #": r["WO #"], "Date": r["Date"], "Total": r["Total"], "Page": r["_Page"], "Complete": r["_Complete"], "Check Score": r["_CheckScore"]} for r in all_rows]
            st.dataframe(audit_rows, use_container_width=True, hide_index=True)
    else:
        render_summary_cards([
            ("Completed WOs", "—", "upload weekly PDF"),
            ("Total Dollars", "—", "waiting for report"),
            ("PDF Rows Reviewed", "—", "waiting for report"),
        ])
        st.info("Upload the weekly PDF to generate the A/R output.")


def render_ap_page():
    render_brand_header()
    st.subheader("A/P")
    st.caption("Best workflow: export the QuickBooks Online A/P Aging Detail report to Excel, then upload it here.")

    ap_upload = st.file_uploader("Upload QuickBooks A/P Aging Detail Excel export", type=["xlsx", "xlsm", "csv", "txt"], key="ap_upload")

    if ap_upload is not None:
        file_bytes = ap_upload.getvalue()
        name = ap_upload.name.lower()
        try:
            if name.endswith((".xlsx", ".xlsm")):
                ap_rows, ap_meta = extract_ap_from_xlsx(file_bytes)
            else:
                text = file_bytes.decode("utf-8-sig", errors="replace")
                ap_rows, ap_meta = parse_ap_csv_text(text)
            ap_payload = {
                "rows": ap_rows,
                "meta": ap_meta,
                "source_label": "uploaded report",
                "filename": ap_upload.name,
                "saved_at": datetime.now().isoformat(),
            }
            st.session_state["ap_report_data"] = ap_payload
            save_persistent_report("ap_report", ap_payload)
            st.success("A/P report cleaned successfully.")
        except Exception as exc:
            st.error("The A/P file could not be processed. Export the A/P Aging Detail report to Excel from QuickBooks and upload the Excel file.")
            st.exception(exc)

    ap_data = st.session_state.get("ap_report_data")

    if ap_data:
        ap_rows = ap_data.get("rows", [])
        ap_meta = ap_data.get("meta", {"company": "White Line Trucking LLC", "report": "A/P Aging Detail Report", "as_of": ""})
        source_label = ap_data.get("source_label", "uploaded report")
        if ap_upload is None:
            st.info(f"Showing last uploaded A/P report: {ap_data.get('filename', 'saved report')}")
    else:
        ap_rows, ap_meta = demo_ap_rows()
        source_label = "sample preview"
        st.info("Showing a sample preview using the layout from the QuickBooks report. Upload the real Excel export to replace this sample.")

    if ap_rows:
        total_ap = sum(money_to_float(r.get("Open balance", "")) for r in ap_rows)
        render_summary_cards([
            ("A/P Rows", f"{len(ap_rows)}", "group headers removed"),
            ("Open Balance Total", f"${total_ap:,.2f}", "clean payable total"),
            ("Source", source_label.title(), "uploaded or preview"),
        ])

        render_ap_report(ap_rows, ap_meta)

        st.subheader("Copy A/P Output")
        ap_tsv_without_due_date = build_tsv(ap_rows, AP_COPY_NO_DUE_COLUMNS, include_header=False, include_total=True)
        ap_tsv_due_date_only = build_tsv(ap_rows, AP_COPY_DUE_DATE_COLUMNS, include_header=False, include_total=False)

        render_ap_two_copy_buttons(ap_tsv_without_due_date, ap_tsv_due_date_only)
    else:
        st.warning("No A/P bill rows were found. The OUT OF RANGE group header and total rows are intentionally removed.")


initialize_persistent_report_state()

page = render_sidebar_nav()

if page == "Home":
    render_home_page()
elif page == "Cash Flow":
    render_cash_flow_page()
elif page == "A/R":
    render_ar_page()
elif page == "A/P":
    render_ap_page()
elif page == "P&L CM":
    render_pl_page("P&L CM", "pl_cm_upload")
elif page == "P&L YTD":
    render_pl_page("P&L YTD", "pl_ytd_upload")
